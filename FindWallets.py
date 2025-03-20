#!/usr/bin/env python3
import requests
import time
import os
import json
import logging
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime, timezone, timedelta

# Funkcja do ładowania konfiguracji z pliku config.json
def load_json_config(config_file="config.json"):
    """
    Ładuje konfigurację z pliku config.json. Jeśli plik nie istnieje, wyświetla komunikat
    i zwraca pusty słownik.
    """
    try:
        with open(config_file, "r") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"BRAK PLIKU {config_file}")
    except Exception as e:
        print(f"Wystąpił błąd podczas ładowania pliku {config_file}: {e}")
    return {}

# ================================ KONSTANTY ================================

# Ładowanie konfiguracji z pliku config.json
config = load_json_config()

# Pobieranie wartości z konfiguracji lub używanie domyślnych
NETWORK = config.get("NETWORK", "ETH")
T1_STR = config.get("T1_STR", "17-03-2025 22:25:00")
T2_STR = config.get("T2_STR", "18-03-2025 19:30:00")
T3_STR = config.get("T3_STR", "19-03-2025 20:00:00")
TOKEN_CONTRACT_ADDRESS = config.get("TOKEN_CONTRACT_ADDRESS", "0x712f43B21cf3e1B189c27678C0f551c08c01D150")

# Dla każdej sieci oddzielny API_KEY oraz API_URL
# BSC
API_KEY_BSC = "A98VM42SB2U2I21QH3HU4CI821YMTYZYWJ"
API_URL_BSC = "https://api.bscscan.com/api"
# ETH
API_KEY_ETH = "ZWPZR1RE3327TU3PWRFSWNF4CM621AZFI9"
API_URL_ETH = "https://api.etherscan.io/api"
# BASE
API_KEY_BASE = "YJA2NW17BI66Y3JV5UVY4FC2BHKXICH7YX"
API_URL_BASE = "https://api.basescan.org/api"

# Wybieramy odpowiednie API_KEY oraz API_URL na podstawie wybranej sieci
if NETWORK == "BSC":
    API_KEY_USED = API_KEY_BSC
    API_URL = API_URL_BSC
elif NETWORK == "ETH":
    API_KEY_USED = API_KEY_ETH
    API_URL = API_URL_ETH
elif NETWORK == "BASE":
    API_KEY_USED = API_KEY_BASE
    API_URL = API_URL_BASE
else:
    raise Exception(f"Nieobsługiwana sieć: {NETWORK}")

# Stałe dotyczące pozostałych ustawień
BLOCK_CHUNK_SIZE = 1200
FREQUENCY_INTERVAL_SECONDS = 60
MIN_FREQ_VIOLATIONS = 5
MIN_TX_COUNT = 10
DELAY_BETWEEN_REQUESTS = 0.2
MAX_RETRIES = 3
MIN_USD_VALUE = 100.0  # Minimalna wartość w USD
WALLETS_FOLDER = "Wallets"
LOGS_FOLDER = "Logs"
CACHE_FILE = os.path.join(WALLETS_FOLDER, "wallet_frequency_cache.json")
DEX_API_URL = "https://api.dexscreener.com/latest/dex/tokens/{}"

# Adres natywnego tokena: dla BSC był WBNB, dla ETH i BASE będzie WETH – tutaj osobno definiujemy
if NETWORK == "BASE":
    BASE_NATIVE_ADDRESS = "0x4200000000000000000000000000000000000006"
    WETH_ADDRESS = BASE_NATIVE_ADDRESS  # BASE używa tego samego adresu co WETH
    NATIVE_TOKEN_NAME = "ETH"
    NATIVE_TOKEN_FULL_NAME = "ethereum"
elif NETWORK == "ETH":
    WETH_ADDRESS = "0xC02aaA39b223FE8D0A0e5C4F27eAD9083C756Cc2"
    NATIVE_TOKEN_NAME = "ETH"
    NATIVE_TOKEN_FULL_NAME = "ethereum"
else:  # BSC
    WBNB_ADDRESS = "0xbb4cdb9cbd36b01bd1cbaebf2de08d9173bc095c"
    NATIVE_TOKEN_NAME = "BNB"
    NATIVE_TOKEN_FULL_NAME = "binancecoin"

LOG_FILE = os.path.join(LOGS_FOLDER, "error_log.txt")

# ================================ UTWORZENIE FOLDERÓW ================================
os.makedirs(WALLETS_FOLDER, exist_ok=True)
os.makedirs(LOGS_FOLDER, exist_ok=True)

# =============================================================================
# Funkcja pomocnicza do podziału zakresu bloków na paczki
def divide_blocks_into_chunks(start_block, end_block, chunk_size):
    chunks = []
    current_start = start_block
    while current_start <= end_block:
        current_end = min(current_start + chunk_size - 1, end_block)
        chunks.append((current_start, current_end))
        current_start = current_end + 1
    return chunks
# =============================================================================

# Konfiguracja logowania – logujemy tylko błędy do jednego pliku
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def parse_date(date_str):
    """
    Konwertuje datę w formacie "DD-MM-YYYY H:M:S" do znacznika unixowego
    i odejmuje 1 godzinę.
    """
    try:
        dt = datetime.strptime(date_str, "%d-%m-%Y %H:%M:%S")
        dt = dt - timedelta(hours=1)  # Odejmujemy 1 godzinę
        dt = dt.replace(tzinfo=timezone.utc)
        return int(dt.timestamp())
    except Exception as e:
        logging.error(f"Błąd parsowania daty {date_str}: {e}")
        raise e

def api_request(params):
    """
    Wykonuje zapytanie do wybranego API (API_URL) z podanymi parametrami.
    Implementuje retry z delay w przypadku błędów oraz walidację formatu odpowiedzi.
    """
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = requests.get(API_URL, params=params, timeout=10)
            if response.status_code == 200:
                data = response.json()
                # Podstawowa walidacja odpowiedzi – sprawdzamy, czy mamy "result"
                if "result" in data:
                    if data.get("status") == "1" or isinstance(data["result"], list):
                        return data
                    elif data.get("message") == "No transactions found":
                        return {"result": []}
                    else:
                        logging.error(f"Błąd API (nieoczekiwany format): {data}")
                else:
                    logging.error(f"Nieoczekiwana struktura odpowiedzi: {data}")
            else:
                logging.error(f"Błąd HTTP {response.status_code}: {response.text}")
        except Exception as e:
            logging.error(f"Wyjątek przy zapytaniu do API: {e}")
        time.sleep(DELAY_BETWEEN_REQUESTS * attempt)
    raise Exception("Nie udało się uzyskać poprawnej odpowiedzi z API po maksymalnej liczbie prób.")

def get_block_by_timestamp(timestamp, closest="before"):
    """
    Pobiera numer bloku na podstawie znacznika czasu.
    """
    params = {
        "module": "block",
        "action": "getblocknobytime",
        "timestamp": timestamp,
        "closest": closest,
        "apikey": API_KEY_USED
    }
    data = api_request(params)
    try:
        block_number = int(data["result"])
        return block_number
    except Exception as e:
        logging.error(f"Nie udało się pobrać numeru bloku dla timestamp {timestamp}: {e}")
        raise

def get_token_transactions(startblock, endblock):
    """
    Pobiera transakcje tokena z zakresu bloków startblock-endblock.
    """
    all_txs = []
    current_start = startblock
    while current_start <= endblock:
        current_end = min(current_start + BLOCK_CHUNK_SIZE - 1, endblock)
        params = {
            "module": "account",
            "action": "tokentx",
            "contractaddress": TOKEN_CONTRACT_ADDRESS,
            "startblock": current_start,
            "endblock": current_end,
            "sort": "asc",
            "apikey": API_KEY_USED
        }
        print(f"Pobieram transakcje dla bloków {current_start} - {current_end}...")
        try:
            data = api_request(params)
            if "result" in data and isinstance(data["result"], list):
                txs = data["result"]
                print(f"Liczba transakcji w odpowiedzi: {len(txs)}")
                all_txs.extend(txs)
            else:
                logging.error(f"Niepoprawny format odpowiedzi API dla bloków {current_start}-{current_end}: {data}")
        except Exception as e:
            logging.error(f"Błąd podczas przetwarzania danych dla bloków {current_start}-{current_end}: {e}")
        finally:
            current_start = current_end + 1
            time.sleep(DELAY_BETWEEN_REQUESTS)
    return all_txs

def load_frequency_cache():
    """
    Ładuje cache weryfikacji częstotliwości transakcji z pliku.
    """
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r") as f:
                return json.load(f)
        except Exception as e:
            logging.error(f"Błąd ładowania cache z {CACHE_FILE}: {e}")
    return {}

def save_frequency_cache(cache):
    """
    Zapisuje cache weryfikacji częstotliwości do pliku.
    """
    try:
        with open(CACHE_FILE, "w") as f:
            json.dump(cache, f)
    except Exception as e:
        logging.error(f"Błąd zapisu cache do {CACHE_FILE}: {e}")

def get_exchange_rate(token_address, retries=5):
    """
    Pobiera kurs wymiany tokena względem natywnego tokena (WBNB dla BSC, WETH dla ETH/BASE) z API Dexscreener.
    """
    attempt = 0
    rate = None
    while attempt < retries:
        try:
            url = DEX_API_URL.format(token_address.lower())
            response = requests.get(url, timeout=10)
            
            if response.status_code != 200:
                logging.error(f"Odpowiedź serwera dla tokena: {response.status_code}, treść: {response.text}")
                raise Exception(f"HTTP error dla tokena: {response.status_code}")
            
            data = response.json()
            pairs = data.get("pairs", [])
            for pair in pairs:
                native_addr = WETH_ADDRESS.lower() if NETWORK in ["ETH", "BASE"] else WBNB_ADDRESS.lower()
                if native_addr in (pair.get("baseToken", {}).get("address", "").lower(),
                                   pair.get("quoteToken", {}).get("address", "").lower()):
                    if pair.get("baseToken", {}).get("address", "").lower() == token_address.lower():
                        rate = float(pair["priceNative"])
                    elif pair.get("quoteToken", {}).get("address", "").lower() == token_address.lower():
                        rate = 1 / float(pair["priceNative"])
                    break
            
            if rate is not None:
                break
            else:
                logging.error(f"Nie znaleziono pary z natywnym tokenem dla tokena {token_address}")
                raise Exception("Brak odpowiedniej pary w danych API")
        
        except Exception as e:
            logging.error(f"Próba {attempt + 1} pobrania kursu nie powiodła się: {e}")
            attempt += 1
            time.sleep(DELAY_BETWEEN_REQUESTS * attempt)
    
    if rate is None:
        rate = "error"
    
    return rate

def get_native_to_usd_rate(retries=3):
    """
    Pobiera kurs wymiany natywnego tokena (ETH, BNB) do USD.
    """
    attempt = 0
    rate = None
    while attempt < retries:
        try:
            url = f"https://api.coingecko.com/api/v3/simple/price?ids={NATIVE_TOKEN_FULL_NAME}&vs_currencies=usd"
            response = requests.get(url, timeout=10)
            
            if response.status_code == 429:
                logging.error("Przekroczono limit zapytań do CoinGecko. Oczekiwanie na kolejną próbę...")
                time.sleep(10)
                attempt += 1
                continue
            
            if response.status_code != 200:
                logging.error(f"Odpowiedź serwera dla natywnego tokena: {response.status_code}, treść: {response.text}")
                raise Exception(f"HTTP error dla natywnego tokena: {response.status_code}")
            
            data = response.json()
            rate = data.get(NATIVE_TOKEN_FULL_NAME, {}).get("usd")
            if rate is not None:
                break
            else:
                logging.error(f"Nie znaleziono kursu wymiany dla natywnego tokena {NATIVE_TOKEN_FULL_NAME}")
                raise Exception("Brak kursu wymiany w danych API")
        
        except Exception as e:
            logging.error(f"Próba {attempt + 1} pobrania kursu natywnego tokena do USD nie powiodła się: {e}")
            attempt += 1
            time.sleep(DELAY_BETWEEN_REQUESTS * attempt)
    
    if rate is None:
        rate = "error"
    
    return rate

def get_last_wallet_transactions(wallet, retries=3, count=10):
    """
    Pobiera ostatnie 'count' transakcji dla danego portfela, niezależnie od tokena,
    korzystając z endpointu txlist.
    """
    attempt = 0
    while attempt < retries:
        try:
            params = {
                "module": "account",
                "action": "txlist",
                "address": wallet,
                "page": 1,
                "offset": count,
                "sort": "desc",
                "apikey": API_KEY_USED
            }
            response = requests.get(API_URL, params=params, timeout=10)
            if response.status_code == 200:
                data = response.json()
                if "result" in data and isinstance(data["result"], list):
                    return data["result"]
                else:
                    logging.error(f"Nieoczekiwany format odpowiedzi przy pobieraniu transakcji portfela {wallet}: {data}")
            else:
                logging.error(f"Błąd HTTP {response.status_code} przy pobieraniu transakcji portfela {wallet}")
        except Exception as e:
            logging.error(f"Wyjątek przy pobieraniu transakcji portfela {wallet}, próba {attempt+1}: {e}")
        attempt += 1
        time.sleep(DELAY_BETWEEN_REQUESTS * attempt)
    return []

def frequency_check(wallet, wallet_txs, cache):
    """
    Sprawdza, czy portfel nie wykonuje transakcji zbyt często.
    Używamy ostatnich 10 transakcji (dotyczących danego tokena).
    """    
    if len(wallet_txs) < MIN_TX_COUNT:
        return True  # Portfel akceptowany, brak wystarczających danych do odrzucenia

    sorted_txs = sorted(wallet_txs, key=lambda x: int(x["timeStamp"]), reverse=True)
    last_10 = sorted_txs[:10]
    
    violations = 0
    for i in range(len(last_10) - 1):
        t1 = int(last_10[i]["timeStamp"])
        t2 = int(last_10[i + 1]["timeStamp"])
        if (t1 - t2) < FREQUENCY_INTERVAL_SECONDS:
            violations += 1
    
    if violations >= MIN_FREQ_VIOLATIONS:
        return False  # Portfel odrzucony
    
    return True  # Portfel akceptowany

def frequency_check_wallet(wallet, cache):
    """
    Dodatkowa weryfikacja portfela – pobieramy ostatnie 10 transakcji (wszystkie typy)
    i sprawdzamy, czy odstępy między transakcjami nie są zbyt krótkie.
    """
    last_txs = get_last_wallet_transactions(wallet, retries=MAX_RETRIES, count=10)
    if not last_txs or len(last_txs) < 2:
        return True  # Brak wystarczających danych do odrzucenia

    violations = 0
    # Sortujemy transakcje od najnowszej do najstarszej
    last_txs_sorted = sorted(last_txs, key=lambda x: int(x["timeStamp"]), reverse=True)
    for i in range(len(last_txs_sorted) - 1):
        t1 = int(last_txs_sorted[i]["timeStamp"])
        t2 = int(last_txs_sorted[i + 1]["timeStamp"])
        if (t1 - t2) < FREQUENCY_INTERVAL_SECONDS:
            violations += 1

    if violations >= MIN_FREQ_VIOLATIONS:
        return False

    return True

def simulate_wallet_balance(wallet, wallet_txs, t1_unix, t2_unix, t3_unix):
    """
    Symuluje saldo tokenów portfela od T1 do T3.
    """
    purchased = 0.0
    balance = 0.0
    purchase_count = 0
    sale_count = 0

    for tx in wallet_txs:
        ts = int(tx["timeStamp"])
        try:
            token_decimals = int(tx.get("tokenDecimal", "0"))
            amount = float(tx["value"]) / (10 ** token_decimals)
        except Exception as e:
            logging.error(f"Błąd przeliczania wartości transakcji: {tx} - {e}")
            continue

        if t1_unix <= ts <= t3_unix:
            if tx["to"].lower() == wallet.lower():
                balance += amount
                if t1_unix <= ts <= t2_unix:
                    purchased += amount
                    purchase_count += 1
            elif tx["from"].lower() == wallet.lower():
                balance -= amount
                sale_count += 1

    return purchased, balance, purchase_count, sale_count

def get_output_filename():
    """
    Tworzy nazwę pliku wynikowego Excel w folderze Wallets.
    """
    base_name = os.path.join(WALLETS_FOLDER, f"{TOKEN_CONTRACT_ADDRESS}.xlsx")
    if not os.path.exists(base_name):
        return base_name
    suffix = 1
    while True:
        new_name = os.path.join(WALLETS_FOLDER, f"{TOKEN_CONTRACT_ADDRESS}_{suffix}.xlsx")
        if not os.path.exists(new_name):
            return new_name
        suffix += 1

def write_excel(filename, header_lines, rows):
    """
    Zapisuje wyniki do pliku Excel (.xlsx) z formatowaniem.
    """

    wb = Workbook()
    ws = wb.active
    current_row = 1
    
    for header in header_lines:
        cell = ws.cell(row=current_row, column=1, value=header)
        cell.font = Font(italic=True, color="808080")
        current_row += 1
    
    current_row += 1  # pusta linia
    
    if rows:
        fieldnames = list(rows[0].keys())
        for col, name in enumerate(fieldnames, start=1):
            cell = ws.cell(row=current_row, column=col, value=name.upper())
            cell.font = Font(bold=True)
        current_row += 1
        
        for row in rows:
            for col, key in enumerate(fieldnames, start=1):
                value = row.get(key, "")
                if key in ["purchased", "final_balance", "native_value", "usd_value"]:
                    try:
                        value = float(value)
                    except:
                        pass
                ws.cell(row=current_row, column=col, value=value)
            current_row += 1
        
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
    else:
        ws.cell(row=current_row, column=1, value="Brak danych")
    
    wb.save(filename)

# Na początku skryptu
start_time = time.time()  # Zapisujemy czas startu

def main():
    try:
        print(f"Wybrana sieć: {NETWORK}")
        print("Rozpoczynam działanie skryptu...")
        t1_unix = parse_date(T1_STR)
        t2_unix = parse_date(T2_STR)
        t3_unix = parse_date(T3_STR)
        print(f"T1: {t1_unix}, T2: {t2_unix}, T3: {t3_unix}")
        
        start_block = get_block_by_timestamp(t1_unix, closest="after")
        end_block = get_block_by_timestamp(t3_unix, closest="before")
        print(f"Zakres bloków: {start_block} - {end_block}")
        
        all_transactions = get_token_transactions(start_block, end_block)
        print(f"Pobrano łącznie {len(all_transactions)} transakcji tokena.")
        
        txs_in_period = [tx for tx in all_transactions if t1_unix <= int(tx["timeStamp"]) <= t3_unix]
        txs_in_period.sort(key=lambda tx: int(tx["timeStamp"]))
        print(f"Transakcje w okresie T1-T3: {len(txs_in_period)}")

        # Wyznaczamy kandydatów na podstawie transakcji tokena – portfele, które wykonały zakup w okresie T1-T2
        candidate_wallets = []
        wallet_transactions = {}
        for tx in txs_in_period:
            wallet_from = tx["from"].lower()
            wallet_to = tx["to"].lower()
            tx_timestamp = int(tx["timeStamp"])
            
            if wallet_from not in wallet_transactions:
                wallet_transactions[wallet_from] = []
            if wallet_to not in wallet_transactions:
                wallet_transactions[wallet_to] = []
            wallet_transactions[wallet_from].append(tx)
            wallet_transactions[wallet_to].append(tx)
            
            if t1_unix <= tx_timestamp <= t2_unix and wallet_to not in candidate_wallets:
                candidate_wallets.append(wallet_to)

        print(f"Znaleziono {len(candidate_wallets)} kandydatów (portfeli z zakupem w okresie T1-T2).")
        
        # Krok 1: Weryfikacja na podstawie transakcji tokena
        frequency_cache = load_frequency_cache()
        filtered_wallets = []
        total_wallets = len(candidate_wallets)
        for index, wallet in enumerate(candidate_wallets, start=1):
            print(f"{index}/{total_wallets}: {wallet}")
            txs = wallet_transactions.get(wallet, [])
            if wallet in frequency_cache:
                print(f"Portfel {wallet} odrzucony (był w cache).")
                continue
            if not frequency_check(wallet, txs, frequency_cache):
                print(f"Portfel {wallet} odrzucony (częste transakcje tokena).")
                continue
            # Krok 2: Dodatkowa weryfikacja – pobieramy ostatnie 10 transakcji (wszystkie transakcje)
            if not frequency_check_wallet(wallet, frequency_cache):
                print(f"Portfel {wallet} odrzucony (częste transakcje pełne).")
                continue
            filtered_wallets.append(wallet)
        print(f"Portfeli po weryfikacji: {len(filtered_wallets)}")
        
        exchange_rate = get_exchange_rate(TOKEN_CONTRACT_ADDRESS, retries=5)
        if exchange_rate == "error":
            print("Nie udało się pobrać kursu wymiany tokena. Wartość natywna ustawiona jako 'error'.")
        else:
            print(f"Kurs wymiany tokena -> {NATIVE_TOKEN_NAME} na dzień T3: {exchange_rate}")
        
        native_to_usd_rate = get_native_to_usd_rate()
        if native_to_usd_rate == "error":
            print("Nie udało się pobrać kursu wymiany natywnego tokena do USD.")
        else:
            print(f"Kurs wymiany {NATIVE_TOKEN_NAME} -> USD: {native_to_usd_rate}")
        
        final_results = []
        for wallet in filtered_wallets:
            txs = wallet_transactions.get(wallet, [])
            purchased, final_balance, purchase_count, sale_count = simulate_wallet_balance(wallet, txs, t1_unix, t2_unix, t3_unix)
            if purchased == 0:
                continue
            percentage = (final_balance / purchased) * 100
            if final_balance < 0.5 * purchased:
                continue
            
            if exchange_rate != "error" and native_to_usd_rate != "error":
                native_value = final_balance * exchange_rate
                usd_value = native_value * native_to_usd_rate
            else:
                native_value = "error"
                usd_value = "error"
            
            if usd_value != "error" and usd_value < MIN_USD_VALUE:
                print(f"Portfel {wallet} odrzucony z powodu niskiej wartości w USD: {usd_value}")
                continue
            
            final_results.append({
                "wallet": wallet,
                "purchase_count": purchase_count,
                "sale_count": sale_count,
                "percentage": f"{percentage:.2f}%",
                "native_value": native_value,
                "usd_value": usd_value,
                "purchased": purchased,
                "final_balance": final_balance,
            })

        save_frequency_cache(frequency_cache)
        print(f"Portfeli po filtracji: {len(final_results)}")
        
        header_lines = [
            f"TOKEN_CONTRACT_ADDRESS: {TOKEN_CONTRACT_ADDRESS}",
            f"T1: {T1_STR}",
            f"T2: {T2_STR}",
            f"T3: {T3_STR}",
            f"NETWORK: {NETWORK}"
        ]
        
        output_filename = get_output_filename()
        write_excel(output_filename, header_lines, final_results)
        
        # Mierzenie czasu do momentu tworzenia pliku
        elapsed_time = time.time() - start_time
        elapsed_minutes, elapsed_seconds = divmod(elapsed_time, 60)
        print(f"Czas wykonania skryptu do momentu zapisu pliku: {int(elapsed_minutes):02}:{int(elapsed_seconds):02}")
    except Exception as e:
        logging.error(f"Błąd głównej funkcji: {e}")
        print("Wystąpił krytyczny błąd. Sprawdź logi w pliku:", LOG_FILE)

if __name__ == "__main__":
    main()

"""
Klasa odpowiedzialna za generowanie raportów Excel.
"""
import os
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font
from config_manager import ConfigManager


class ExcelReporter:
    """Generuje raporty Excel z wynikami analizy portfeli"""
    
    def __init__(self, config_manager: ConfigManager):
        self.config_manager = config_manager
        paths = config_manager.get_paths_config()
        self.wallets_folder = paths["wallets_folder"]
        
        # Zapewniamy istnienie folderu
        os.makedirs(self.wallets_folder, exist_ok=True)
    
    def _get_unique_filename(self, token_address: str) -> str:
        """Generuje unikalną nazwę pliku dla raportu"""
        base_name = os.path.join(self.wallets_folder, f"{token_address}.xlsx")
        
        if not os.path.exists(base_name):
            return base_name
        
        # Jeśli plik już istnieje, dodajemy sufiks
        suffix = 1
        while True:
            new_name = os.path.join(self.wallets_folder, f"{token_address}_{suffix}.xlsx")
            if not os.path.exists(new_name):
                return new_name
            suffix += 1
    
    def _create_header_info(self) -> List[str]:
        """Tworzy informacje nagłówkowe dla raportu"""
        config = self.config_manager.config
        
        return [
            f"TOKEN_CONTRACT_ADDRESS: {config.get('TOKEN_CONTRACT_ADDRESS', 'N/A')}",
            f"T1: {config.get('T1_STR', 'N/A')}",
            f"T2: {config.get('T2_STR', 'N/A')}",
            f"T3: {config.get('T3_STR', 'N/A')}",
            f"NETWORK: {config.get('NETWORK', 'N/A')}"
        ]
    
    def _format_cell_value(self, value: Any, column_key: str) -> Any:
        """Formatuje wartość komórki w zależności od typu kolumny"""
        if column_key in ["purchased", "final_balance", "native_value", "usd_value"]:
            try:
                return float(value) if value != "error" else value
            except (ValueError, TypeError):
                return value
        
        return value
    
    def _auto_adjust_column_width(self, worksheet) -> None:
        """Automatycznie dopasowuje szerokość kolumn"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            
            # Ustawiamy szerokość z małym marginesem
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    def generate_report(self, results: List[Dict[str, Any]], token_address: str) -> str:
        """
        Generuje raport Excel z wynikami analizy.
        Zwraca ścieżkę do utworzonego pliku.
        """
        filename = self._get_unique_filename(token_address)
        
        # Tworzenie nowego arkusza
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Wallet Analysis"
        
        current_row = 1
        
        # Dodawanie informacji nagłówkowych
        header_lines = self._create_header_info()
        for header_line in header_lines:
            cell = worksheet.cell(row=current_row, column=1, value=header_line)
            cell.font = Font(italic=True, color="808080")
            current_row += 1
        
        # Pusta linia po nagłówkach
        current_row += 1
        
        if results:
            # Nagłówki kolumn
            column_headers = list(results[0].keys())
            for col_idx, header in enumerate(column_headers, start=1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=header.upper())
                cell.font = Font(bold=True)
            
            current_row += 1
            
            # Dane wyników
            for result in results:
                for col_idx, key in enumerate(column_headers, start=1):
                    value = result.get(key, "")
                    formatted_value = self._format_cell_value(value, key)
                    worksheet.cell(row=current_row, column=col_idx, value=formatted_value)
                
                current_row += 1
        else:
            # Brak danych
            cell = worksheet.cell(row=current_row, column=1, value="Brak danych do wyświetlenia")
            cell.font = Font(italic=True, color="FF0000")
        
        # Automatyczne dopasowanie szerokości kolumn
        self._auto_adjust_column_width(worksheet)
        
        # Zapisanie pliku
        try:
            workbook.save(filename)
            print(f"Report saved to: {filename}")
            return filename
        except Exception as e:
            raise Exception(f"Error saving Excel report: {e}")
    
    def generate_summary_statistics(self, results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Generuje statystyki podsumowujące dla analizy"""
        if not results:
            return {
                "total_wallets": 0,
                "total_purchased": 0,
                "total_final_balance": 0,
                "average_retention": 0,
                "total_usd_value": 0
            }
        
        total_wallets = len(results)
        total_purchased = 0
        total_final_balance = 0
        total_usd_value = 0
        valid_usd_count = 0
        
        for result in results:
            try:
                purchased = float(result.get("purchased", 0))
                final_balance = float(result.get("final_balance", 0))
                total_purchased += purchased
                total_final_balance += final_balance
                
                # Zliczamy tylko ważne wartości USD
                usd_value = result.get("usd_value", 0)
                if usd_value != "error" and isinstance(usd_value, (int, float)):
                    total_usd_value += float(usd_value)
                    valid_usd_count += 1
                    
            except (ValueError, TypeError):
                continue
        
        average_retention = (
            (total_final_balance / total_purchased * 100) 
            if total_purchased > 0 else 0
        )
        
        return {
            "total_wallets": total_wallets,
            "total_purchased": round(total_purchased, 2),
            "total_final_balance": round(total_final_balance, 2),
            "average_retention": f"{average_retention:.2f}%",
            "total_usd_value": round(total_usd_value, 2),
            "wallets_with_usd_value": valid_usd_count
        }
    
    def print_summary(self, results: List[Dict[str, Any]], execution_time: str = None) -> None:
        """Wyświetla podsumowanie wyników analizy"""
        stats = self.generate_summary_statistics(results)
        
        print("\n" + "="*50)
        print("PODSUMOWANIE ANALIZY")
        print("="*50)
        print(f"Liczba znalezionych portfeli: {stats['total_wallets']}")
        print(f"Łączna ilość zakupionych tokenów: {stats['total_purchased']}")
        print(f"Łączna ilość pozostałych tokenów: {stats['total_final_balance']}")
        print(f"Średni poziom retencji: {stats['average_retention']}")
        
        if stats['wallets_with_usd_value'] > 0:
            print(f"Łączna wartość USD: ${stats['total_usd_value']}")
            print(f"Portfeli z wartością USD: {stats['wallets_with_usd_value']}")
        
        if execution_time:
            print(f"Czas wykonania: {execution_time}")
        
        print("="*50)
import os
import re
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font
from .config_manager import ConfigManager
from shared.constants.config_constants import ConfigConstants

class ExcelReporter:

    def __init__(self, config_manager: ConfigManager):
        self.config_manager = config_manager
        paths = config_manager.get_paths_config()
        self.wallets_folder = paths["wallets_folder"]

        os.makedirs(self.wallets_folder, exist_ok=True)

    def _get_unique_filename(self, token_name: str, t1_str: str, t2_str: str, t3_str: str) -> str:

        def format_date_for_filename(date_str):
            try:
                from datetime import datetime
                dt = datetime.strptime(date_str, ConfigConstants.DATE_FORMAT)
                return dt.strftime("%d-%m-%Y")
            except ValueError:
                return date_str.replace(" ", "_").replace(":", "").replace("-", "")

        t1_formatted = format_date_for_filename(t1_str)
        t2_formatted = format_date_for_filename(t2_str)
        t3_formatted = format_date_for_filename(t3_str)

        safe_token_name = re.sub(r'[\\/:*?"<>|]', "_", token_name.replace("$", ""))

        base_name = os.path.join(self.wallets_folder, f"{safe_token_name}__T1_{t1_formatted}__T2_{t2_formatted}__T3_{t3_formatted}.xlsx")

        if not os.path.exists(base_name):
            return base_name

        suffix = 1
        while True:
            new_name = os.path.join(self.wallets_folder, f"{safe_token_name}__T1_{t1_formatted}__T2_{t2_formatted}__T3_{t3_formatted}__{suffix}.xlsx")
            if not os.path.exists(new_name):
                return new_name
            suffix += 1

    def _format_cell_value(self, value: Any, column_key: str) -> Any:

        if column_key in ["purchased", "final_balance", "native_value", "usd_value"]:
            if value is None:
                return "N/A"
            try:
                return float(value)
            except (ValueError, TypeError):
                return value

        return value

    def _auto_adjust_column_width(self, worksheet) -> None:

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)

            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    def generate_report(self, results: List[Dict[str, Any]], token_name: str, t1_str: str, t2_str: str, t3_str: str) -> str:

        filename = self._get_unique_filename(token_name, t1_str, t2_str, t3_str)

        workbook = Workbook()
        worksheet = workbook.active
        if worksheet is None:
            raise RuntimeError("Failed to create the Excel worksheet")
        worksheet.title = "Wallet Analysis"

        current_row = 1

        config_data = [
            ("", self.config_manager.config.get('TOKEN_CONTRACT_ADDRESS', 'N/A')),
            ("", token_name if token_name != "error" else 'N/A'),
            ("", self.config_manager.config.get('NETWORK', 'N/A')),
            ("", self.config_manager.config.get('T1_STR', 'N/A')),
            ("", self.config_manager.config.get('T2_STR', 'N/A')),
            ("", self.config_manager.config.get('T3_STR', 'N/A'))
        ]

        for param, value in config_data:
            worksheet.cell(row=current_row, column=1, value=param)
            worksheet.cell(row=current_row, column=2, value=value)
            current_row += 1

        current_row += 2

        if results:

            cell = worksheet.cell(row=current_row, column=1, value="NR")
            cell.font = Font(bold=True)

            column_headers = list(results[0].keys())
            for col_idx, header in enumerate(column_headers, start=2):
                cell = worksheet.cell(row=current_row, column=col_idx, value=header.upper())
                cell.font = Font(bold=True)

            current_row += 1

            for row_num, result in enumerate(results, start=1):

                worksheet.cell(row=current_row, column=1, value=row_num)

                for col_idx, key in enumerate(column_headers, start=2):
                    value = result.get(key, "")
                    formatted_value = self._format_cell_value(value, key)
                    worksheet.cell(row=current_row, column=col_idx, value=formatted_value)

                current_row += 1
        else:

            cell = worksheet.cell(row=current_row, column=1, value="Brak danych")
            cell.font = Font(italic=True, color="FF0000")

        if results:
            self._auto_adjust_column_width(worksheet)

        worksheet.column_dimensions['A'].width = 5
        worksheet.column_dimensions['B'].width = 60

        try:
            workbook.save(filename)
            return filename
        except Exception as e:
            raise Exception(f"Error saving Excel report: {e}")
